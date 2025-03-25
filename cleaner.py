import pandas as pd
import re
import openpyxl
import os
from openpyxl import load_workbook

def clean_excel(input_file, output_file):
    """
    First phase: Clean and extract relevant data from the input Excel file.
    
    This function:
    1. Standardizes column names
    2. Extracts MODELNO., QTY./CASE, and PRICE columns
    3. Cleans price values, removing commas and extracting numeric values
    4. Removes '*' from MODELNO. column values
    5. Adds CASE, Price, and Code columns
    """
    print(f"Starting first phase: Cleaning data from {input_file}")
    
    # Load the Excel file
    xls = pd.ExcelFile(input_file)
    
    # Dictionary to store cleaned data
    cleaned_sheets = {}

    for sheet_name in xls.sheet_names:
        # Read each sheet
        df = xls.parse(sheet_name)
        
        # If it's the "GuidedNav_184360-Cleaned" sheet, apply special transformations
        if sheet_name == "GuidedNav_184360-Cleaned":
            # Remove column 0
            df.drop(df.columns[0], axis=1, inplace=True)

            # Rename headers
            new_headers = [
                "MODELNO.", "INSIDE DIM.L x W x H", "VARIABLE DEPTH/DESCRIPTION",
                "PRICE PER BOX", "PRICE PER BOX2", "PRICE PER BOX3", "PRICE PER BOX4",
                "PRICE PER BOX5", "ADD TOCART", "ADD"
            ]
            df.columns = new_headers  # Assign new column names

            # Remove empty row at index 3 (if it exists)
            if len(df) > 3 and df.iloc[3].isnull().all():
                df.drop(index=3, inplace=True)
                df.reset_index(drop=True, inplace=True)
                
        # Standardize column names (lowercase and stripped of whitespace)
        df.columns = [col.lower().strip() for col in df.columns]

        # Identify columns
        modelno_column = [col for col in df.columns if col == "modelno."]
        qty_column = [col for col in df.columns if col == "qty./case"]
        price_columns = [col for col in df.columns if re.search(r'(^|\s)price($|\s)', col)]

        # Ensure at least one relevant column exists
        if not modelno_column and not price_columns:
            print(f"Skipping sheet '{sheet_name}' - No relevant columns found.")
            continue

        # Arrange columns: MODELNO. -> QTY./CASE (if available) -> PRICE columns
        final_columns = modelno_column + qty_column + price_columns
        cleaned_df = df[final_columns].copy()

        # Remove '*','▪' from MODELNO. column values
        if modelno_column:
            col_name = modelno_column[0]  # Ensure the column exists
            cleaned_df[col_name] = cleaned_df[col_name].astype(str).fillna("").str.replace(r'[\*▪]', '', regex=True)



        # Clean price columns
        for col in price_columns:
            cleaned_df[col] = (
                cleaned_df[col]
                .astype(str)
                .str.replace(',', '', regex=True)  # Remove commas
                .apply(lambda x: re.search(r'\d*\.\d+|\d+', x).group() if re.search(r'\d*\.\d+|\d+', x) else None)
            )
            cleaned_df[col] = pd.to_numeric(cleaned_df[col], errors='coerce')  # Convert valid numeric values

        # Append three additional columns
        cleaned_df["CASE"] = ""
        cleaned_df["Price"] = ""
        cleaned_df["Code"] = ""

        # If QTY./CASE is missing, fill "Price" with combined price column values
        if not qty_column:
            cleaned_df["Price"] = cleaned_df[price_columns].apply(lambda row: ', '.join(row.dropna().astype(str)), axis=1)

        # Store cleaned data
        cleaned_sheets[sheet_name] = cleaned_df

    # Save cleaned data to a new Excel file
    temp_output = "temp_cleaned_output.xlsx"
    with pd.ExcelWriter(temp_output) as writer:
        for sheet, data in cleaned_sheets.items():
            data.to_excel(writer, sheet_name=sheet, index=False)

    print(f"First phase complete: Intermediate file saved as {temp_output}")
    return temp_output

def update_case_and_price_columns(input_file, output_file):
    """
    Second phase: Update CASE and Price columns in the cleaned Excel file.
    
    For sheets with QTY./CASE:
    - CASE column: Multiply QTY./CASE by each price multiplier (e.g., 100 * 1, 100 * 3, etc.)
    - Price column: Calculate QTY./CASE divided by price (e.g., 100/17 = 5.88, 100/16 = 6.25, etc.)
    
    For sheets without QTY./CASE:
    - CASE column: Copy price multipliers directly
    - Price column: Not modified
    """
    print(f"\nStarting second phase: Updating CASE and Price columns in {input_file}")
    
    # Load the workbook
    workbook = load_workbook(input_file)
    
    # Process each sheet
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"Processing sheet: {sheet_name}")
        
        # Find header row and locate necessary columns
        headers = [cell.value for cell in sheet[1]]
        
        # Find CASE column index
        case_idx = None
        for idx, header in enumerate(headers, 1):  # Excel is 1-indexed
            if header == 'CASE':
                case_idx = idx
                break
        
        if not case_idx:
            print(f"  Skipping sheet {sheet_name} - no CASE column found")
            continue
        
        # Find Price column index
        price_col_idx = None
        for idx, header in enumerate(headers, 1):
            if header == 'Price':
                price_col_idx = idx
                break
        
        if not price_col_idx:
            print(f"  Warning: Price column not found in sheet {sheet_name}")
            
        # Find QTY./CASE column index
        qty_idx = None
        for idx, header in enumerate(headers, 1):
            if header and isinstance(header, str) and 'qty' in header.lower():
                qty_idx = idx
                break
        
        # Find price columns
        price_indices = []
        for idx, header in enumerate(headers, 1):
            if header and isinstance(header, str) and 'price' in header.lower() and header != 'Price':
                price_indices.append(idx)
        
        if not price_indices:
            print(f"  Skipping sheet {sheet_name} - no price columns found")
            continue
        
        # Get price multipliers from second row
        second_row = [cell.value for cell in sheet[2]]
        price_multipliers = []
        
        for idx in price_indices:
            value = second_row[idx-1] if idx-1 < len(second_row) else None
            if isinstance(value, (int, float)):
                price_multipliers.append(value)
            elif isinstance(value, str) and value.replace(',', '').replace('.', '', 1).isdigit():
                # For price multipliers, ignore values with commas as per requirement
                if ',' in value:
                    price_multipliers.append(None)
                else:
                    price_multipliers.append(float(value))
            else:
                price_multipliers.append(None)
        
        print(f"  Price multipliers: {price_multipliers}")
        
        # Process data rows
        for row_idx in range(3, sheet.max_row + 1):  # Start from third row (index 3)
            # If we have a QTY./CASE column
            if qty_idx:
                qty_cell = sheet.cell(row=row_idx, column=qty_idx)
                qty_value = qty_cell.value
                
                # Extract quantity as number
                qty = None
                if isinstance(qty_value, (int, float)):
                    qty = qty_value
                elif isinstance(qty_value, str) and qty_value.replace(',', '').replace('.', '', 1).isdigit():
                    # For QTY./CASE, we should handle values with commas
                    qty = float(qty_value.replace(',', ''))
                
                if not qty:
                    continue  # Skip if no valid quantity
                
                # Calculate CASE values and Price values
                case_values = []
                price_values = []
                
                for idx, multiplier in enumerate(price_multipliers):
                    if multiplier is not None:
                        # For CASE column - multiply qty by multiplier
                        case_result = qty * multiplier
                        # Format as integer if it's a whole number
                        case_values.append(str(int(case_result) if case_result == int(case_result) else case_result))
                        
                        # For Price column - get corresponding price per case value
                        if price_col_idx and idx < len(price_indices):
                            price_idx = price_indices[idx]
                            price_value = sheet.cell(row=row_idx, column=price_idx).value
                            
                            # Only process if price value exists and is numeric
                            if price_value is not None:
                                if isinstance(price_value, (int, float)):
                                    # Calculate the division result
                                    division_result = price_value / qty
                                    # Format as integer if it's a whole number
                                    price_values.append(str(int(division_result) if division_result == int(division_result) else round(division_result, 2)))
                                elif isinstance(price_value, str):
                                    # Remove $ and commas for price value conversion
                                    clean_price_str = price_value.replace(',', '').replace('$', '').strip()
                                    if clean_price_str and clean_price_str.replace('.', '', 1).isdigit():
                                        clean_price = float(clean_price_str)
                                        division_result = clean_price / qty 
                                        price_values.append(str(int(division_result) if division_result == int(division_result) else round(division_result, 2)))
                
                # Update CASE cell
                case_cell = sheet.cell(row=row_idx, column=case_idx)
                case_cell.value = ', '.join(case_values)
                
                # Update Price cell if it exists and we have values to update
                if price_col_idx and price_values:
                    price_cell = sheet.cell(row=row_idx, column=price_col_idx)
                    price_cell.value = ', '.join(price_values)
                
                if row_idx <= 5:  # Log first few updates
                    model_no = sheet.cell(row=row_idx, column=1).value
                    print(f"  Updated row {row_idx}: ModelNo={model_no}, QTY={qty}, CASE={case_cell.value}")
                    if price_col_idx and price_values:
                        print(f"  Price column updated: Price={', '.join(price_values)}")
                
            else:
                # For sheets without QTY column, copy multipliers directly for CASE
                # For Price, leave it unchanged as per requirements
                case_values = []
                for multiplier in price_multipliers:
                    if multiplier is not None:
                        case_values.append(str(int(multiplier) if multiplier == int(multiplier) else multiplier))
                
                # Update CASE cell
                case_cell = sheet.cell(row=row_idx, column=case_idx)
                case_cell.value = ', '.join(case_values)
                
                if row_idx <= 5:  # Log first few updates
                    model_no = sheet.cell(row=row_idx, column=1).value
                    print(f"  Updated row {row_idx}: ModelNo={model_no}, CASE={case_cell.value}")
                    # Price not updated for sheets without QTY./CASE
    
    # Save the workbook
    workbook.save(output_file)
    print(f"Second phase complete: Final file saved as {output_file}")

def append_case_price(uline_csv, processed_xlsx, output_csv):
    # Load the CSV file
    uline_df = pd.read_csv(uline_csv)
    
    # Load the Excel file (all sheets)
    processed_df = pd.read_excel(processed_xlsx, sheet_name=None)
    
    # Combine all sheets into one DataFrame
    processed_combined = pd.concat(processed_df.values(), ignore_index=True)
    
    # Ensure relevant columns exist
    required_columns = {'modelno.', 'CASE', 'Price'}
    missing_columns = required_columns - set(processed_combined.columns)
    if missing_columns:
        raise ValueError(f"Missing columns in processed_output.xlsx: {missing_columns}")
    
    if 'manu_sku' not in uline_df.columns:
        raise ValueError("Column 'manu_sku' not found in uline_scrap_products.csv")
    
    # Merge based on 'manu_sku' (CSV) and 'modelno.' (Excel)
    merged_df = uline_df.merge(
        processed_combined[['modelno.', 'CASE', 'Price']],
        left_on='manu_sku',
        right_on='modelno.',
        how='left'
    )

    # Drop redundant 'modelno.' column
    merged_df.drop(columns=['modelno.'], inplace=True)

    # Convert 'CASE' and 'Price' columns to list format
    def to_list(value):
        if pd.isna(value) or value == "":
            return ""  # Keep empty if no value
        return [int(x.strip()) if x.strip().isdigit() else float(x.strip()) for x in str(value).split(',')]

    merged_df['CASE'] = merged_df['CASE'].apply(to_list)
    merged_df['Price'] = merged_df['Price'].apply(to_list)

    # Generate 'Code' column as a list based on the length of 'Price'
    def generate_code(price):
        if price == "" or not isinstance(price, list):
            return ""  # Keep empty if Price is empty
        return ['C'] * len(price)

    merged_df['Code'] = merged_df['Price'].apply(generate_code)

    # Save the updated CSV
    merged_df.to_csv(output_csv, index=False)
    print(f"Updated CSV saved as {output_csv}")

def process_excel_file(input_file, output_file):
    """
    Main function that combines both cleaning and updating processes:
    1. First cleans the data from the original file
    2. Then updates the CASE and Price columns
    """
    print(f"Starting complete Excel processing on {input_file}")
    
    # Phase 1: Clean the data
    temp_output = clean_excel(input_file, output_file)
    
    # Phase 2: Update CASE and Price columns
    update_case_and_price_columns(temp_output, output_file)
    
    print(f"Processing complete. Final file saved as {output_file}")

    if os.path.exists(temp_output):
        os.remove(temp_output)
        print(f"Temporary file {temp_output} has been deleted.")

if __name__ == "__main__":
    # Run the full process on the input file
    process_excel_file('html_output.xlsx', 'processed_output.xlsx')
    append_case_price("uline_scrap_products.csv", "processed_output.xlsx", "updated_uline_scrap_products.csv")
